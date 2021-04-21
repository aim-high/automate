# Description: 	student_app_import.py prepares the student app file for import into FMP.
#				On a rolling basis throughout the summer, paper student applications are
#				entered into Google Forms. This script automates creating and deleting
#				columns.
# Purpose: 		Reformat paper student applications from Google Forms' export in order to
#				import into student database.
# Author: 		Erica Ching (eching@aimhigh.org)
# Created: 		04-12-2021

#!/usr/bin/python3
import openpyxl
from datetime import date

now = date.today()
now = now.strftime('%m-%d-%Y')

# Must update Google file's name when applicable
wb = openpyxl.load_workbook('2021 Student Application - Paper (Responses).xlsx')

# make copy of original export's tab for Person
original = wb['Form Responses 1']
ws = wb.copy_worksheet(original)
ws.title = 'Person'
ws = wb['Person']

# Use excel's COLUMN function to find the corresponding column number
# delete_cols() deletes columns starting from the first argument number PLUS the 2nd argument number
ws.delete_cols(65,12) 		# delete columns 65:76 (alpha is BM:BX)
ws.delete_cols(18,43)		# delete columns 18:60 (alpha is R:BH)
ws.delete_cols(16,1)		# delete column 16
ws.delete_cols(10,5)		# delete columns 10:14
ws.delete_cols(6,3)			# delete columns 6:8
ws.delete_cols(1,3) 		# delete columns 1:3

# Delete rows
ws.delete_rows(5,1)
ws.delete_rows(1,3)

# Insert and set InitialIdentification field for first record in first column
ws.insert_cols(1)
ws['A1'] = 'InitialIdentification'
ws['A2'] = 'Student'

# make copy of original for Student tab

ws = wb.copy_worksheet(original)
ws.title = 'Student app'
ws = wb['Student app']

# Prepare for VLOOKUPs of school region and site choices
# insert_cols: insert a row at 24 (before the existing row 24)

# Insert one column to right of current school region and site choices
ws.insert_cols(24)  # school region
ws.insert_cols(9)   # site choice 3
ws.insert_cols(8)   # site choice 2
ws.insert_cols(7)   # site choice 1

# Delete old headers and Person columns
ws.delete_rows(5,1)
ws.delete_rows(1,3)

"""
ws.delete_cols(61,4)
ws.delete_cols(17,1)
ws.delete_cols(15,1)
ws.delete_cols(9,1)
ws.delete_cols(4,2)
"""

# Insert 5 columns for name combined, ID_Person_student, Year, ApplicationType, ApplicationMode, StatusOfApplication
ws.insert_cols(1,6)
ws['A1'] = 'name combined'
ws['B1'] = 'ID_Person_student'
ws['C1'] = 'Year'
ws['C2'] = 2021
ws['D1'] = 'ApplicationType'
ws['D2'] = 'New'
ws['E1'] = 'ApplicationMode'
ws['E2'] = 'Paper'
ws['F1'] = 'StatusOfApplication'
ws['F2'] = 'Submitted'

wb.save('Google Forms student app import to FMP ' + now + '.xlsx')
print('Renamed file: Google Forms student app import to FMP ' + now + '.xlsx')
