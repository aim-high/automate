# Description: 	student_app_import.py prepares the student app file for import into FMP. 
#				On a rolling basis throughout the summer, paper student applications are 
#				entered into Google Forms. This script automates creating and deleting 
#				columns.
# Purpose: 		Reformat paper student applications from Google Forms' export in order to 
#				import into student database. 
# Author: 		Erica Ching (eching@aimhigh.org)
# Created: 		08-06-2019
# Modified:		03-19-2020

#!/usr/bin/python3
import openpyxl
from datetime import date

now = date.today()
now = now.strftime('%m-%d-%Y')

# Must update Google file's name when applicable
wb = openpyxl.load_workbook('2020 Student Application - General (Responses).xlsx')
ws = wb['Person']

# Use excel's COLUMN function to find the corresponding column number
# delete_cols() deletes columns starting from the first argument number PLUS the 2nd argument number 
ws.delete_cols(77,15) 		# delete columns 77:90 (alpha is BY:CM)
ws.delete_cols(19,54)		# delete columns 19:72 (alpha is S:BT)
ws.delete_cols(16,2)		# delete columns 16:17
ws.delete_cols(11,4)		# delete columns 11:14
ws.delete_cols(6,4)			# delete columns 6:9
ws.delete_cols(1,3)			# delete columns 1:3

# Insert and set InitialIdentification field for first record
ws.insert_cols(1)
ws['A1'] = 'InitialIdentification'
ws['A2'] = 'Student'

ws = wb['Student app']
# Insert one column to right of current school region
ws.insert_cols(22)

# Insert one column to the right of Site choices
ws.insert_cols(10)
ws.insert_cols(9)
ws.insert_cols(8)

# Insert 5 columns for name combined, ID_Person_student, Year, ApplicationType, ApplicationMode, StatusOfApplication
ws.insert_cols(1,6)
ws['A1'] = 'name combined'
ws['B1'] = 'ID_Person_student'
ws['C1'] = 'Year'
ws['C2'] = 2020
ws['D1'] = 'ApplicationType'
ws['D2'] = 'New'
ws['E1'] = 'ApplicationMode'
ws['E2'] = 'Paper'
ws['F1'] = 'StatusOfApplication'
ws['F2'] = 'Submitted'

wb.save('Google Forms student app import to FMP ' + now + '.xlsx')
print('Renamed file: Google Forms student app import to FMP ' + now + '.xlsx')
