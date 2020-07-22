# Description: Deletes daily attendance columns and cleans up other unnecessary 
# columns to import DaysAttended and AttendedOverTenDaysFlag matched with ID_Person_Student
#
# Author: Erica Ching (eching@aimhigh.org)

import openpyxl
from datetime import date

now = date.today()
now = now.strftime('%m-%d-%Y')

wb = openpyxl.load_workbook('2019 Attendance UPA.xlsx')
ws = wb.get_sheet_by_name('6th Grade')
ws.delete_cols(6,27)
ws.delete_cols(2,2)
ws['G1'] = 'DaysAttended'
ws['H1'] = 'AttendedOverTenDaysFlag'
ws['A1'] = 'ID_Person_Student'


ws = wb.get_sheet_by_name('7th Grade')
ws.delete_cols(6,27)
ws.delete_cols(2,2)
ws['G1'] = 'DaysAttended'
ws['H1'] = 'AttendedOverTenDaysFlag'
ws['A1'] = 'ID_Person_Student'

ws = wb.get_sheet_by_name('8th Grade')
ws.delete_cols(6,27)
ws.delete_cols(2,2)
ws['G1'] = 'DaysAttended'
ws['H1'] = 'AttendedOverTenDaysFlag'
ws['A1'] = 'ID_Person_Student'


ws = wb.get_sheet_by_name('9th Grade')
ws.delete_cols(6,27)
ws.delete_cols(2,2)
ws['G1'] = 'DaysAttended'
ws['H1'] = 'AttendedOverTenDaysFlag'
ws['A1'] = 'ID_Person_Student'
# insert one column to right of current school region

#ws.insert_cols(1,5)

wb.save('2019 Attendance UPA ' + now + '.xlsx')
print('Renamed file: 2019 Attendance [Site] ' + now + '.xlsx')
