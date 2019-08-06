# Purpose: In order to import math scores into database, must format data by flipping axis on Excel. 
# Author: Erica Ching (eching@aimhigh.org)

#! python3
# automate.py - read TestQuestionID, StudentID, and TestScore from 'Math Pretest Francisco.xlsx' and copy it into 'test.xlsx' in separate columns 


import openpyxl, pprint
from openpyxl import load_workbook, Workbook

print('Opening workbook...')

# get math pretest
wb1 = load_workbook('Math Pretest Francisco.xlsx')
sheet1 = wb1.worksheets[0]

# write into another sheet
wb2 = load_workbook('test.xlsx')
sheet2 = wb2.worksheets[0]

# grab info from wb1
tuple(sheet1['A1':'X1'])

# grab TestQuestionId
print('Reading from Math Pretest Francisco.xlsx...')
print('~~ ID_TestQuestion ~~')

columnNum=1
rowNum=1
for num in sheet1['A1':'A51']: 
	for rowOfCellObjects in sheet1['G1':'X1']:
		for cellObj in rowOfCellObjects:
			#print(cellObj.coordinate, cellObj.value)
			sheet2.cell(rowNum,columnNum).value = cellObj.value
			rowNum += 1


print('~~ ID_Person_Student ~~')

columnNum=2
rowNum=1
oldCol=1
oldRow=5	
for rowOfCellObjects in sheet1['G5':'X51']:
	for cellObj in rowOfCellObjects:
		#print(cellObj.coordinate, cellObj.value)
		sheet2.cell(rowNum,columnNum).value = sheet1.cell(oldRow,oldCol).value
		rowNum += 1
	oldRow += 1

print('~~ QuestionScore_Pre ~~')
# G5:X5 starts
columnNum=3
rowNum=1

tuple(sheet1['G5':'X51'])
for rowOfCellObjects in sheet1['G5':'X51']:
	for cellObj in rowOfCellObjects:
		#print(cellObj.coordinate, cellObj.value)
		sheet2.cell(rowNum,columnNum).value = cellObj.value	
		rowNum += 1

print('~~ Saving test.xlsx ~~')
# save file
wb2.save('test.xlsx')
